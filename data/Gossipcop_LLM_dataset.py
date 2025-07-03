
import torch.utils.data as data
import torchvision.transforms as transforms
import os
import pandas as pd
import openpyxl
from tqdm import tqdm
import torchvision.transforms.functional as F
from PIL import Image
import cv2
import numpy as np
import torch

class Gossipcop_LLM_dataset(data.Dataset):

    def __init__(
            self, root_path='./dataset/gossipcop_LLM', dataset='gossip',
                 is_filter=True, is_use_unimodal=True, image_size=224, is_train=True,
                 data_augment=False, with_ambiguity=False, use_soft_label=False,is_sample_positive=1,
                 duplicate_fake_times=0,
                 not_on_12=0, downsample_rate=0.5
                 ):
        # not_on_12 = not_on_12 > 0
        # print(f"not on 12? {not_on_12}")
        # self.duplicate_fake_times = duplicate_fake_times
        # self.with_ambiguity = with_ambiguity
        # self.use_soft_label = use_soft_label
        # self.data_augment = data_augment
        self.root_path = root_path
        self.dataset_name = dataset
        super(Gossipcop_LLM_dataset, self).__init__()
        # self.is_sample_positive = is_sample_positive
        self.is_train = is_train
        self.label_dict, self.label_ambiguity = [], []
        self.image_size = image_size
        self.resize_and_to_tensor = transforms.Compose([
                transforms.Resize((224,224)),
                transforms.ToTensor(),]
        )
        self.image_folder = '/root/autodl-tmp/data/top_img'
        workbook_name = self.root_path + '/{}_datasets_gossipcop_LLM.xlsx'.format('train' if self.is_train else 'test')
        wb = openpyxl.load_workbook(workbook_name)
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row
        
        for i in tqdm(range(2, rows + 1)):
            content = str(sheet['B' + str(i)].value)
            image_name = str(sheet['C' + str(i)].value)
            label = int(sheet['D' + str(i)].value)
            news_type = int(sheet['E' + str(i)].value)
            record = {}
            record['images'] = image_name
            record['label'] = label
            record['content'] = content
            record['category'] = news_type
            self.label_dict.append(record)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t

    def __getitem__(self, index):
        GT_size = self.image_size

        record = self.label_dict[index]
        images, label, content, category = record['images'], record['label'], record['content'], record['category']

        if '_' in images:
            images = images.split('_')[0]
        if category == 0:
            image_path = os.path.join(self.image_folder, images+'_top_img.png')
            img_GT = cv2.imread(image_path, cv2.IMREAD_COLOR)
            img_GT = img_GT.astype(np.float32) / 255.
            if img_GT.ndim == 2:
                img_GT = np.expand_dims(img_GT, axis=2)
            # some images have 4 channels
            if img_GT.shape[2] > 3:
                img_GT = img_GT[:, :, :3]

            img_GT = channel_convert(img_GT.shape[2], 'RGB', [img_GT])[0]
            H_origin, W_origin, _ = img_GT.shape

            ###### directly resize instead of crop
            img_GT = cv2.resize(np.copy(img_GT), (GT_size, GT_size),
                                interpolation=cv2.INTER_LINEAR)

            orig_height, orig_width, _ = img_GT.shape
            H, W, _ = img_GT.shape

            # BGR to RGB, HWC to CHW, numpy to tensor
            if img_GT.shape[2] == 3:
                img_GT = img_GT[:, :, [2, 1, 0]]

            img_GT = torch.from_numpy(np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))).float()
        else:
            img_GT = torch.zeros((3,GT_size,GT_size))
        return content, img_GT, label, category, images


def channel_convert(in_c, tar_type, img_list):
    # conversion among BGR, gray and y
    if in_c == 3 and tar_type == 'gray':  # BGR to gray
        gray_list = [cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) for img in img_list]
        return [np.expand_dims(img, axis=2) for img in gray_list]
    elif in_c == 3 and tar_type == 'y':  # BGR to y
        y_list = [bgr2ycbcr(img, only_y=True) for img in img_list]
        return [np.expand_dims(img, axis=2) for img in y_list]
    elif in_c == 1 and tar_type == 'RGB':  # gray/y to BGR
        return [cv2.cvtColor(img, cv2.COLOR_GRAY2BGR) for img in img_list]
    else:
        return img_list

def bgr2ycbcr(img, only_y=True):
    '''bgr version of rgb2ycbcr
    only_y: only return Y channel
    Input:
        uint8, [0, 255]
        float, [0, 1]
    '''
    in_img_type = img.dtype
    img.astype(np.float32)
    if in_img_type != np.uint8:
        img *= 255.
    # convert
    if only_y:
        rlt = np.dot(img, [24.966, 128.553, 65.481]) / 255.0 + 16.0
    else:
        rlt = np.matmul(img, [[24.966, 112.0, -18.214], [128.553, -74.203, -93.786],
                              [65.481, -37.797, 112.0]]) / 255.0 + [16, 128, 128]
    if in_img_type == np.uint8:
        rlt = rlt.round()
    else:
        rlt /= 255.
    return rlt.astype(in_img_type)



